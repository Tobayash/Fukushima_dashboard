from __future__ import annotations

from pathlib import Path
import re
from urllib.request import Request, urlopen
import zipfile

import openpyxl
import pandas as pd

from extractors.common import AREAS, EVACUATION_12, RAW_DIR, area_group, empty_records_df, normalize_area_id, now_iso


SUICIDE_SOURCE_NAME = "福島県 自殺関連指標を計算するためのエクセルシート"
SUICIDE_SOURCE_PAGE_URL = "https://www.pref.fukushima.lg.jp/sec/21840a/s-statistics-ctv.html"
SUICIDE_ZIP_URL = "https://www.pref.fukushima.lg.jp/uploaded/attachment/738226.zip"
SUICIDE_ZIP_PATH = RAW_DIR / "suicide_statistics_202604.zip"
SUICIDE_EXTRACT_DIR = RAW_DIR / "suicide_statistics_202604"
RATE_MIN_POPULATION = 10000


def _ensure_suicide_workbooks() -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    if not SUICIDE_ZIP_PATH.exists():
        request = Request(SUICIDE_ZIP_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(request, timeout=60) as response:
            SUICIDE_ZIP_PATH.write_bytes(response.read())

    needed = [
        SUICIDE_EXTRACT_DIR / "suicide_jinkoudoutai.xlsx",
        SUICIDE_EXTRACT_DIR / "ctv_jinkou.xlsx",
    ]
    if not all(path.exists() for path in needed):
        SUICIDE_EXTRACT_DIR.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(SUICIDE_ZIP_PATH) as archive:
            archive.extractall(SUICIDE_EXTRACT_DIR)
    return SUICIDE_EXTRACT_DIR


def _sheet_year(sheet_name: str) -> int | None:
    heisei = re.fullmatch(r"H(\d{2})", sheet_name)
    if heisei:
        return 1988 + int(heisei.group(1))
    reiwa = re.fullmatch(r"W(\d{2})", sheet_name)
    if reiwa:
        return 2018 + int(reiwa.group(1))
    return None


def _target_year_sheets(workbook: openpyxl.Workbook) -> list[tuple[str, int]]:
    sheets: list[tuple[str, int]] = []
    for sheet_name in workbook.sheetnames:
        year = _sheet_year(sheet_name)
        if year is not None and 2010 <= year <= 2026:
            sheets.append((sheet_name, year))
    return sorted(sheets, key=lambda row: row[1])


def _population_by_area(workbook_dir: Path) -> dict[tuple[str, int], float]:
    workbook = openpyxl.load_workbook(workbook_dir / "ctv_jinkou.xlsx", data_only=True, read_only=True)
    values: dict[tuple[str, int], float] = {}
    for sheet_name, year in _target_year_sheets(workbook):
        ws = workbook[sheet_name]
        for row in ws.iter_rows(min_row=1, values_only=True):
            area_name = row[1] if len(row) > 1 else None
            population = row[2] if len(row) > 2 else None
            if area_name not in EVACUATION_12 or population in (None, 0):
                continue
            try:
                values[(str(area_name), year)] = float(population)
            except (TypeError, ValueError):
                continue
    return values


def _suicide_deaths_by_area(workbook_dir: Path) -> dict[tuple[str, int], float]:
    workbook = openpyxl.load_workbook(workbook_dir / "suicide_jinkoudoutai.xlsx", data_only=True, read_only=True)
    values: dict[tuple[str, int], float] = {}
    for sheet_name, year in _target_year_sheets(workbook):
        ws = workbook[sheet_name]
        for row in ws.iter_rows(min_row=1, values_only=True):
            label = str(row[0] or "").strip()
            area_name = re.sub(r"^\d+", "", label).strip().replace("\u3000", "")
            if area_name not in EVACUATION_12:
                continue
            deaths = row[1] if len(row) > 1 else None
            if deaths is None:
                deaths = 0
            values[(str(area_name), year)] = float(deaths)
    return values


def _record(
    *,
    indicator_id: str,
    indicator_name: str,
    area_name: str,
    period: str,
    value: float | None,
    unit: str,
    notes: str,
    retrieved_at: str,
) -> dict[str, object]:
    return {
        "indicator_id": indicator_id,
        "indicator_name": indicator_name,
        "category": "身体的・精神的健康",
        "concept": "精神的健康・自殺関連指標",
        "area_id": normalize_area_id(area_name),
        "area_name": area_name,
        "area_group": area_group(area_name),
        "period": period if value is not None else "",
        "value": value,
        "unit": unit,
        "source_type": "行政統計",
        "source_name": SUICIDE_SOURCE_NAME,
        "source_url": SUICIDE_SOURCE_PAGE_URL,
        "retrieved_at": retrieved_at,
        "collection_method": "B",
        "notes": notes,
    }


def fetch_suicide_health_indicators() -> pd.DataFrame:
    """自殺関連指標を福島県公式Excelから市町村別・年次で取得する。"""
    rows: list[dict[str, object]] = []
    retrieved_at = now_iso()
    try:
        workbook_dir = _ensure_suicide_workbooks()
        deaths = _suicide_deaths_by_area(workbook_dir)
        populations = _population_by_area(workbook_dir)
    except Exception as exc:  # noqa: BLE001 - keep NO DATA rows visible.
        notes = f"NO DATA: 福島県公式Excelから自殺関連指標を取得できませんでした（{exc}）。"
        for area in AREAS:
            if area["area_id"] == "pref_fukushima":
                continue
            for indicator_id, indicator_name, unit in [
                ("suicide_deaths_vital", "自殺者数（人口動態統計）", "人"),
                ("suicide_rate_vital", "自殺死亡率（人口動態統計・人口10万対）", "人/10万人"),
            ]:
                rows.append(
                    _record(
                        indicator_id=indicator_id,
                        indicator_name=indicator_name,
                        area_name=area["area_name"],
                        period="",
                        value=None,
                        unit=unit,
                        notes=notes,
                        retrieved_at=retrieved_at,
                    )
                )
        return pd.concat([empty_records_df(), pd.DataFrame(rows)], ignore_index=True)

    for area in AREAS:
        if area["area_id"] == "pref_fukushima":
            continue
        area_name = area["area_name"]
        area_years = sorted(year for name, year in deaths if name == area_name)
        if not area_years:
            notes = "NO DATA: 福島県公式Excelに当該市町村の人口動態統計ベース自殺者数を確認できませんでした。"
            for indicator_id, indicator_name, unit in [
                ("suicide_deaths_vital", "自殺者数（人口動態統計）", "人"),
                ("suicide_rate_vital", "自殺死亡率（人口動態統計・人口10万対）", "人/10万人"),
            ]:
                rows.append(
                    _record(
                        indicator_id=indicator_id,
                        indicator_name=indicator_name,
                        area_name=area_name,
                        period="",
                        value=None,
                        unit=unit,
                        notes=notes,
                        retrieved_at=retrieved_at,
                    )
                )
            continue

        for year in area_years:
            period = f"{year}-12-31"
            death_count = deaths[(area_name, year)]
            population = populations.get((area_name, year))
            rate = death_count / population * 100000 if population and population >= RATE_MIN_POPULATION else None
            base_note = (
                "人口動態統計に基づく市町村別自殺者数。人口は同ZIP内の福島県人口推計を使用。"
                f"小規模自治体では1人の増減で率が大きく変動するため、分母人口が{RATE_MIN_POPULATION:,}人未満の年は死亡率をNO DATAにしています。"
            )
            rows.append(
                _record(
                    indicator_id="suicide_deaths_vital",
                    indicator_name="自殺者数（人口動態統計）",
                    area_name=area_name,
                    period=period,
                    value=death_count,
                    unit="人",
                    notes=base_note,
                    retrieved_at=retrieved_at,
                )
            )
            rows.append(
                _record(
                    indicator_id="suicide_rate_vital",
                    indicator_name="自殺死亡率（人口動態統計・人口10万対）",
                    area_name=area_name,
                    period=period,
                    value=rate,
                    unit="人/10万人",
                    notes=base_note,
                    retrieved_at=retrieved_at,
                )
            )

    return pd.concat([empty_records_df(), pd.DataFrame(rows)], ignore_index=True)
